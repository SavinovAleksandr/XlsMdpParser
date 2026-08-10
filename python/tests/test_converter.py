from pathlib import Path
import shutil

from mdp_converter.core import convert_directory
from mdp_converter.diagnostics import ParseDiagnostics
from mdp_converter.expression import parse_expression
from mdp_converter.expression.evaluator import (
    active_branch,
    collect_variables,
    evaluate,
    format_expression,
)
from mdp_converter.html_generator import generate
from mdp_converter.normalization import factor_key, split_numbered, strip_formula_prefix
from mdp_converter.parameter_analysis import analyze_parameters
from mdp_converter.parse_pipeline import parse
from mdp_converter.models import RowData, Scheme
from mdp_converter.table_detector import control_columns, detect_columns, header_map


ROOT = Path(__file__).resolve().parents[1]
SOURCES = ROOT / "Исходные файлы"


def test_trailing_expression_footnotes_are_stripped_before_parsing():
    assert strip_formula_prefix("2) 320-Pотб 1") == "320-Pотб"
    assert strip_formula_prefix("310-Pотб-Pнб2 2") == "310-Pотб-Pнб2"
    assert strip_formula_prefix("320-Pотб - 1") == "320-Pотб - 1"


def test_nested_if_with_comma_arguments_and_double_equals():
    ast = parse_expression("if(g>=3,if(r==1,313+0.29*x,303+0.2*x),245)")
    env = {"g": 3, "r": 1, "x": 10}
    assert evaluate(ast, env) == 315.9
    assert "IF" not in format_expression(active_branch(ast, env))


def test_russian_decimal_comma_and_semicolon_arguments():
    ast = parse_expression("IF(g>=3;313+0,29*x;245)")
    assert evaluate(ast, {"g": 3, "x": 10}) == 315.9


def test_generated_identifiers_accept_spaces_number_sign_and_ampersand():
    ast = parse_expression(
        "if(S_АОПО_Сясь__–_Колп_№1_2__S22109_==1 & "
        "S_4_ст__АОПО_Сясь__–_Колп_№1_2__S22110_==1, "
        "500-Рнеб блок КАЭС__I18134_, 9999)"
    )
    env = {
        name: (25 if name.startswith("Рнеб") else 1)
        for name in collect_variables(ast)
    }
    assert evaluate(ast, env) == 475


def test_factor_identity_ignores_excel_identifier_formatting():
    assert factor_key("Рнеб блок КАЭС [I18134]") == factor_key(
        "Рнеб_блок_КАЭС__I18134_"
    )
    assert factor_key("S АОПО Сясь* – Колп №1,2 [S22109]") == factor_key(
        "S_АОПО_Сясь__–_Колп_№1_2__S22109_"
    )
    assert factor_key("Режим (северный)") == factor_key("Режим__северный_")


def test_diagnostics_do_not_repeat_identical_messages():
    diag = ParseDiagnostics()
    diag.warn("Одинаковое предупреждение")
    diag.warn("Одинаковое предупреждение")
    diag.error("Одинаковая ошибка")
    diag.error("Одинаковая ошибка")
    assert diag.warnings == ["Одинаковое предупреждение"]
    assert diag.errors == ["Одинаковая ошибка"]


def test_case_is_return_expression_becomes_selectable_mode_parameter():
    ast = parse_expression(
        "case(Сезон_АОПО_В–Я, "
        "is(1), return (200), "
        "is(4), return (245), "
        "is(2,3), return (270))"
    )
    variable = "Сезон_АОПО_В_Я"
    assert evaluate(ast, {variable: 1}) == 200
    assert evaluate(ast, {variable: 4}) == 245
    assert evaluate(ast, {variable: 2}) == 270
    assert evaluate(ast, {variable: 3}) == 270
    assert evaluate(ast, {variable: 5}) == 9999

    mode_params, factors = analyze_parameters([ast], [])
    assert not factors
    assert len(mode_params) == 1
    assert mode_params[0].name == variable
    assert mode_params[0].kind == "select"
    assert {option.value for option in mode_params[0].options} == {"1", "2", "3", "4"}
    assert mode_params[0].default == "1"
    assert {option.value: option.label for option in mode_params[0].options}["1"] == "Группа 1"


def test_case_return_may_contain_formula_and_have_arithmetic_suffix():
    ast = parse_expression(
        "case(Сезон_АОПО, is(1), return (280), "
        "is(0), return (300+Pотб))+0.5*Pарх"
    )
    assert evaluate(ast, {"Сезон_АОПО": 1, "Pотб": 20, "Pарх": 100}) == 330
    assert evaluate(ast, {"Сезон_АОПО": 0, "Pотб": 20, "Pарх": 100}) == 370
    mode_params, _ = analyze_parameters([ast], [])
    assert [(param.name, param.kind) for param in mode_params] == [
        ("Сезон_АОПО", "select")
    ]
    assert format_expression(
        active_branch(ast, {"Сезон_АОПО": 1, "Pотб": 20, "Pарх": 100})
    ) == "280 + 0.5 × Pарх"


def test_mode_parameters_are_not_duplicated_as_editable_factors():
    ast = parse_expression(
        "case(Сезон_АОПО_В–Я, is(1), return(200), is(2,3), return(270))-Pотб+Рон"
    )
    modes, factors = analyze_parameters(
        [ast],
        [
            "Сезон АОПО (Вологодская – Явенга)",
            "Pотб",
            "Рон",
            "Неиспользуемый фактор",
        ],
    )
    assert [mode.name for mode in modes] == ["Сезон_АОПО_В_Я"]
    assert [factor.name for factor in factors] == ["Pотб", "Рон"]


def test_criteria_columns_stop_before_control_group():
    # Minimal three-level header used by the real workbooks.
    rows = [
        ["", "", "", "МДП без ПА", "МДП с ПА", "АДП", "Критерий определения", "", "", "", "Контроль дополнительных параметров"],
        ["", "", "", "", "", "", "", "", "", "", "МДП без ПА"],
        ["", "", "", "", "", "", "МДП без ПА", "", "МДП с ПА", "АДП"],
    ]
    hm = header_map(rows)
    assert hm["criteria_mdp"][0] == 6
    assert hm["criteria_mdp_pa"][0] == 8
    assert hm["criteria_adp"][0] == 9


def test_control_group_is_mapped_to_logical_parameter_columns():
    rows = [
        ["№", "Схема сети", "ТНВ", "МДП без ПА", "МДП с ПА", "АДП", "Контроль дополнительных параметров", "", "", "АДП", "Примечание"],
        ["", "", "", "", "", "", "МДП без ПА", "МДП с ПА1", "МДП с ПА2", "АДП", ""],
        ["1", "Нормальная схема", "0", "100", "110", "200", "ДТН-1", "АОПО-1", "АОПО-2", "ДДТН-1", ""],
    ]
    hm = header_map(rows)
    columns = control_columns(rows, hm, True, True)
    assert columns == {"mdp": [6], "mdp_pa": [7, 8], "adp": [9]}


def test_all_supplied_workbooks_parse_without_formula_warnings():
    files = sorted(p for p in SOURCES.glob("*.xlsx") if not p.name.startswith("~$"))
    assert len(files) >= 17
    for source in files:
        diag = ParseDiagnostics()
        model = parse(source, diag)
        assert model.schemes, source.name
        assert model.title != "Контролируемое сечение", source.name
        formula_warnings = [w for w in diag.warnings if w.startswith("Формула не разобрана")]
        assert not formula_warnings, f"{source.name}: {formula_warnings[:3]}"


def test_information_sheet_metadata_is_parsed_for_all_workbooks():
    expected_oscillations = {4.0, 5.0, 10.0, 25.0, 40.0}
    files = sorted(p for p in SOURCES.glob("*.xlsx") if not p.name.startswith("~$"))
    parsed = [parse(source) for source in files]
    assert {model.irregular_oscillation_mw for model in parsed} == expected_oscillations

    pechora = parse(SOURCES / "КС Печорская ГРЭС - Ухта.xlsx")
    assert pechora.irregular_oscillation_mw == 25.0
    assert pechora.weather_stations == ["Печорская ГРЭС", "ПС 220 кВ Зеленоборск"]

    kv = parse(SOURCES / "К-В.xlsx")
    assert kv.irregular_oscillation_mw == 25.0
    assert kv.weather_stations == []

    ukhta_mikun = parse(SOURCES / "КС Ухта-Микунь.xlsx")
    assert ukhta_mikun.weather_stations == [
        "ПС 220 кВ Микунь",
        "ПС 220 кВ Синдор",
        "ПС 220 кВ Сыктывкар",
        "ПС 220 кВ Ухта",
    ]


def test_batch_conversion_processes_folder_and_continues_after_failure(tmp_path):
    source_dir = tmp_path / "xlsx"
    output_dir = tmp_path / "html"
    source_dir.mkdir()
    shutil.copy2(SOURCES / "КС Печорская ГРЭС - Ухта.xlsx", source_dir / "01.xlsx")
    shutil.copy2(SOURCES / "Сосногорск.xlsx", source_dir / "02.XLSX")
    (source_dir / "03.xlsx").write_text("не является книгой Excel", encoding="utf-8")
    (source_dir / "~$временный.xlsx").write_text("служебный файл", encoding="utf-8")
    (source_dir / "прочитать.txt").write_text("не Excel", encoding="utf-8")

    converted, failures = convert_directory(source_dir, output_dir, True, False)

    assert [path.name for path in converted] == ["01.html", "02.html"]
    assert [path.name for path, _ in failures] == ["03.xlsx"]
    assert all(path.exists() for path in converted)
    assert not (output_dir / "~$временный.html").exists()


def test_uncontrolled_status_is_explicit_and_not_inferred_from_empty_rows():
    va = parse(SOURCES / "В-А.xlsx")
    uncontrolled = next(s for s in va.schemes if s.number == "18")
    assert uncontrolled.is_controlled is False
    assert not uncontrolled.rows
    assert all(s.is_controlled for s in va.schemes if s.number != "18")


def test_condition_column_is_used_as_temperature_category():
    model = parse(SOURCES / "ПГРЭС-Инта ВР.xlsx")
    first = model.schemes[0]
    assert len(first.rows) == 6
    assert first.rows[0].temperature == "-20"
    assert len(first.rows[0].mdp_items) == 2


def test_threshold_mode_offers_values_on_both_sides():
    model = parse(SOURCES / "КС Печорская ГРЭС - Ухта.xlsx")
    count = next(p for p in model.mode_params if p.name == "Кол_во_ТГ_ПГРЭС")
    assert {o.value for o in count.options} >= {"2", "3"}


def test_temperature_groups_collect_leading_criteria_before_temperature_cell():
    model = parse(SOURCES / "КС Печорская ГРЭС - Ухта.xlsx")
    scheme = next(item for item in model.schemes if item.number == "9")
    assert [row.temperature for row in scheme.rows] == [
        "-5 и менее", "0", "5", "10", "15", "20", "25", "30", "35", "40 и более"
    ]
    for row in scheme.rows[:-1]:
        assert [item.number for item in row.mdp_items] == [1, 2, 3, 4, 5]
        assert [item["number"] for item in split_numbered(row.crit_mdp)] == [1, 2, 3, 4, 5]


def test_temperature_dependent_schemes_have_no_orphan_rows_or_duplicate_criteria():
    model = parse(SOURCES / "КС Печорская ГРЭС - Ухта.xlsx")
    for scheme in model.schemes:
        if any(row.temperature for row in scheme.rows):
            assert all(row.temperature for row in scheme.rows), scheme.number
        for row in scheme.rows:
            numbers = [item.number for item in row.mdp_items]
            assert len(numbers) == len(set(numbers)), (scheme.number, row.temperature, numbers)


def test_technical_pa_placeholders_do_not_enable_pa_group():
    model = parse(SOURCES / "Сосногорск.xlsx")
    assert model.has_mdp is True
    assert model.has_mdp_pa is False
    assert model.has_adp is True


def test_only_pa_group_is_detected_without_mdp_group():
    rows = [
        ["№", "Схема сети", "ТНВ", "МДП с ПА", "АДП"],
        ["1", "Нормальная схема", "0", "1) 150", "180"],
    ]
    hm = header_map(rows)
    assert detect_columns(rows, hm) == (False, True, True)


def test_adp_criteria_are_read_from_actual_populated_column():
    for name in ("Сосногорск.xlsx", "КС Печорская ГРЭС - Ухта.xlsx"):
        model = parse(SOURCES / name)
        pairs = [
            (row.adp, row.crit_adp)
            for scheme in model.schemes
            for row in scheme.rows
            if row.adp
        ]
        assert pairs, name
        assert all(criteria for _, criteria in pairs), name


def test_named_automation_columns_are_combined_as_mdp_with_pa():
    for name in ("В-А.xlsx", "Микунь - Сыктывкар.xlsx"):
        model = parse(SOURCES / name)
        assert model.has_mdp_pa is True, name
        pa_rows = [row for scheme in model.schemes for row in scheme.rows if row.mdp_pa]
        assert pa_rows and any(row.mdp_pa_items for row in pa_rows), name
        assert any(row.crit_mdp_pa for row in pa_rows), name


def test_va_adp_criteria_do_not_include_pa_automation_criteria():
    model = parse(SOURCES / "В-А.xlsx")
    criteria = {
        row.crit_adp
        for scheme in model.schemes
        for row in scheme.rows
        if row.crit_adp
    }
    assert criteria
    assert all("Исключение срабатывания на ДС АОПО" not in value for value in criteria)
    uncontrolled = next(s for s in model.schemes if s.number == "18")
    assert not uncontrolled.rows
    assert "не контролируется" in uncontrolled.note.lower()


def test_konosha_arpm_uses_setpoint_group_axis():
    candidates = [
        Path("/Users/asavinov/Downloads/Исходные/Коноша - Вельск АРПМ.xlsx"),
        SOURCES / "Коноша - Вельск АРПМ.xlsx",
    ]
    path = next((candidate for candidate in candidates if candidate.exists()), None)
    if path is None:
        import pytest

        pytest.skip("Konosha ARPM workbook is not available")
    model = parse(path)
    assert model.row_axis_label == "Группа уставок АРПМ"
    rows = model.schemes[0].rows
    assert len(rows) == 4
    assert [row.temperature for row in rows] == ["1", "2", "3", "4"]
    assert all(len(row.mdp_pa_items) == 1 for row in rows)
    assert all(row.crit_mdp_pa.startswith("1)") for row in rows)


def test_adp_criteria_skip_minimum_highlight_in_html(tmp_path):
    model = parse(SOURCES / "К-В.xlsx")
    output = tmp_path / "konosha.html"
    generate(model, output)
    text = output.read_text(encoding="utf-8")
    assert "criteriaBlock(items, env, critText, highlightMin=true)" in text
    assert "adpCriteria,false)" in text
    assert "function withMinimumPrefix(html, rawText)" in text
    assert '<div class="minimum-from">Минимальный из:</div>' in text


def test_mikun_syktyvkar_adp_columns_use_rowspan_with_grid_borders(tmp_path):
    candidates = [
        Path("/Users/asavinov/Downloads/Исходные/Микунь – Сыктывкар (под новые МУ).xlsx"),
        SOURCES / "Микунь - Сыктывкар.xlsx",
    ]
    path = next((candidate for candidate in candidates if candidate.exists()), None)
    if path is None:
        import pytest

        pytest.skip("Mikun-Syktyvkar workbook is not available")
    model = parse(path)
    output = tmp_path / "mikun.html"
    generate(model, output)
    text = output.read_text(encoding="utf-8")
    assert "function mergedCellAttrs(count, className)" in text
    assert 'rowspan="${count}" class="${className}"' in text
    assert "border-collapse:collapse" in text
    assert "vertical-align:top" in text
    assert "function paintMergedCellDividers()" not in text
    assert "function rowAdpBlock(row, env)" not in text


def test_vologda_scheme_control_columns_merge_across_temperatures(tmp_path):
    candidates = [
        Path("/Users/asavinov/Downloads/Исходные/Вологда - Архангельск.xlsx"),
        SOURCES / "В-А.xlsx",
    ]
    path = next((candidate for candidate in candidates if candidate.exists()), None)
    if path is None:
        import pytest

        pytest.skip("Vologda workbook is not available")
    model = parse(path)
    scheme = next(s for s in model.schemes if s.number == "1")
    assert scheme.rows[0].control_adp
    assert all(not row.control_adp for row in scheme.rows[1:])
    output = tmp_path / "va.html"
    generate(model, output)
    text = output.read_text(encoding="utf-8")
    assert "function controlValuesVary(rows,key)" in text
    assert 'mergedCellAttrs(rowCount,\'control-cell\')' in text


def test_vologda_arkhangelsk_uses_aopo_setpoint_labels():
    model = parse(SOURCES / "В-А.xlsx")
    assert model.has_pa_seasons is True
    assert model.pa_season_label == "Группа уставок АОПО"
    labels = {option.value: option.label for option in model.pa_season_options}
    assert labels == {
        "1": "Летняя уставка",
        "2": "Зимняя уставка",
        "3": "Весенне-осенняя уставка",
    }
    pa_param = next(p for p in model.mode_params if p.name == "pa_season")
    assert {option.value: option.label for option in pa_param.options} == labels
    variant_labels = {
        variant.label
        for scheme in model.schemes
        for row in scheme.rows
        for variant in row.mdp_pa_variants
    }
    assert "Летняя уставка" in variant_labels
    assert "Зимняя уставка" in variant_labels
    assert "Весенне-осенняя уставка" in variant_labels


def test_conditional_adp_adds_generator_mode_parameter():
    model = parse(SOURCES / "ПГРЭС-Инта.xlsx")
    param = next(p for p in model.mode_params if p.name == "количество_генераторов_в_работе_на_ВТЭЦ_2")
    assert param.kind == "select"
    assert {option.value for option in param.options} == {"1", "2", "3", "4"}
    assert any(row.adp_items for scheme in model.schemes for row in scheme.rows if row.adp)


def test_formula_qualified_factors_replace_unused_generic_names():
    model = parse(SOURCES / "К-В.xlsx")
    names = {factor.name for factor in model.factors}
    assert "Pнб" not in names
    assert "Pон" not in names
    assert "Pнб__Коноша___Вельск__Р" in names
    assert "Pон__Коноша___Вельск__Р" in names
    assert len({name for name in names if name.startswith(("Pнб", "Pон"))}) == 2


def test_reference_ui_shell_and_no_calculation_variant(tmp_path):
    model = parse(SOURCES / "КС Печорская ГРЭС - Ухта.xlsx")
    full = tmp_path / "full.html"
    plain = tmp_path / "plain.html"
    generate(model, full, True, True)
    generate(model, plain, False, False)
    full_text = full.read_text(encoding="utf-8")
    plain_text = plain.read_text(encoding="utf-8")
    assert 'class="hero"' in full_text
    assert 'class="top-grid"' in full_text
    assert "Раскрыть все" in full_text and "Свернуть все" in full_text
    assert "Показано: ${shown} из ${DATA.schemes.length}" in full_text
    assert "График зависимости МДП" in full_text
    assert 'id="chartTooltip"' in full_text
    assert "Перемещайте указатель по графику" in full_text
    assert "schemeAdpBlock" in full_text
    assert 'id="calcToggle"' in full_text
    assert "Вычислять МДП" in full_text
    assert "String(o.value)===String(p.default)" in full_text
    assert "inline-value" in full_text
    assert "mdp-total" not in full_text
    assert "copyMdpSeries" in full_text
    assert "Копировать МДП без ПА" in full_text
    assert "Нерегулярные колебания" in full_text
    assert "Печорская ГРЭС" in full_text
    assert ".elements li::before{content:'⚡'" in full_text

    va = parse(SOURCES / "В-А.xlsx")
    va_html = tmp_path / "va.html"
    generate(va, va_html, True, True)
    va_text = va_html.read_text(encoding="utf-8")
    assert "uncontrolled-badge" in va_text
    assert '"is_controlled": false' in va_text
    assert "function padCopyColumn(value,targetStop)" in full_text
    assert "function formatCopyFactorDefinitions(usedNames)" in full_text
    assert "function collectSchemeFormulaVariables(scheme, group)" in full_text
    assert "Math.ceil((targetStop-length)/8)" in full_text
    assert "Состояние реактора на Печорской ГРЭС" in full_text
    assert "border-left:1px solid #dfe6ef" in full_text
    assert ".criteria-start{border-left:1px solid #dfe6ef}" in full_text
    assert "const formulaHead=" in va_text
    assert "const criteriaHead=" in va_text
    assert "'<th>МДП без ПА</th>'" in va_text or "'<th>МДП без ПА</th>':''" in va_text
    assert "Критерии МДП без ПА</th>" in va_text
    assert "criteria-start" in va_text
    assert ".adp-cell .formula .value{font-size:11px;font-weight:400;color:#1e293b" in full_text
    assert ".col-formula{width:30%}.col-criterion{width:38%}" in full_text
    assert "График зависимости МДП" not in plain_text
    assert "Влияющие факторы для расчёта МДП" not in plain_text
    assert '<details class="card"><summary class="card-h">Влияющие факторы</summary>' in plain_text
    assert '<details class="card"><summary class="card-h">Влияющие факторы</summary>' in full_text
    assert '<details class="card"><summary class="card-h">Значения влияющих факторов</summary>' in full_text
    assert "Значения влияющих факторов" not in plain_text
    assert '</ul><div class="info-grid">' in plain_text
    assert "фактическое значение наибольшего возможного небаланса" in plain_text
    assert 'id="calcToggle"' not in plain_text
    assert '<div class="card-h">Режим расчёта</div>' in plain_text
    assert "renderModeControls();" in plain_text
    assert "Контроль доп. параметров АДП" in full_text
    assert "planning-badge" in full_text
    assert "Для планирования" in full_text


def test_ddtn_criteria_backfills_missing_control_from_scheme_value():
    scheme = Scheme(number="3", name="Test")
    scheme.rows = [
        RowData(
            temperature="25 и менее",
            crit_mdp="1) АДТН\n2) 10% Р",
            control_mdp="ДДТН ВЛ 220 кВ Печорская ГРЭС - Зеленоборск",
        ),
        RowData(
            temperature="30",
            crit_mdp="1) АДТН\n2) 10% Р\n3) ДДТН ВЛ 220 кВ Печорская ГРЭС – Ухта",
            control_mdp="",
        ),
    ]
    from mdp_converter.parse_pipeline import _consolidate_scheme_controls, _ensure_control_for_ddtn_criteria

    _consolidate_scheme_controls(scheme)
    _ensure_control_for_ddtn_criteria(scheme)
    assert scheme.rows[0].control_mdp == "ДДТН ВЛ 220 кВ Печорская ГРЭС - Зеленоборск"
    assert scheme.rows[1].control_mdp == ""


def test_ddtn_criteria_infers_control_once_when_scheme_value_missing():
    scheme = Scheme(number="25", name="Test")
    scheme.rows = [
        RowData(
            temperature="30",
            crit_mdp="1) АДТН\n3) ДДТН ВЛ 110 кВ Чикшино – Каджером",
            control_mdp="",
        ),
        RowData(
            temperature="35",
            crit_mdp="1) АДТН\n3) ДДТН ВЛ 110 кВ Чикшино – Каджером",
            control_mdp="",
        ),
    ]
    from mdp_converter.parse_pipeline import _ensure_control_for_ddtn_criteria

    _ensure_control_for_ddtn_criteria(scheme)
    assert scheme.rows[0].control_mdp == "ДДТН ВЛ 110 кВ Чикшино – Каджером"
    assert scheme.rows[1].control_mdp == ""


def test_pechora_vr_scheme_with_ddtn_keeps_control_once_per_scheme():
    candidates = [
        Path("/Users/asavinov/Downloads/Исходные/Печорская ГРЭС – Ухта ВР (новые МУ_кор ВФ).xlsx"),
        SOURCES / "КС Печорская ГРЭС - Ухта (ВР).xlsx",
    ]
    path = next((candidate for candidate in candidates if candidate.exists()), None)
    if path is None:
        import pytest

        pytest.skip("Pechora VR workbook is not available")
    scheme = next(s for s in parse(path).schemes if s.number == "3")
    rows_with_ddtn = [
        row
        for row in scheme.rows
        if "ддтн" in row.crit_mdp.lower()
    ]
    filled = [row for row in scheme.rows if row.control_mdp]
    assert rows_with_ddtn
    assert len(filled) == 1
    assert filled[0].control_mdp


def test_copy_footer_uses_factor_definitions_from_info_sheet(tmp_path):
    candidates = [
        Path("/Users/asavinov/Downloads/Исходные/Микунь – Урдома (актуализация_кор ВФ).xlsx"),
        SOURCES / "КС Микунь-Урдома.xlsx",
    ]
    path = next((candidate for candidate in candidates if candidate.exists()), None)
    if path is None:
        import pytest

        pytest.skip("Mikun-Urdoma workbook is not available")
    model = parse(path)
    names = {factor.name for factor in model.factor_definitions}
    assert "PУрд" in names
    assert any("Рнб" in name for name in names)
    descriptions = {factor.name: factor.description for factor in model.factor_definitions}
    assert "Переток активной мощности по 1АТ" in descriptions["PУрд"]
    output = tmp_path / "urdoma.html"
    generate(model, output, True, False)
    text = output.read_text(encoding="utf-8")
    assert "formatCopyFactorDefinitions(usedNames)" in text
    assert "collectSchemeFormulaVariables(scheme,group)" in text
    assert "copyNeedsMinimumPrefix(rawText, prepared.length)" in text
    assert "lines.push('где:')" in text


def test_real_control_values_and_planning_markers_are_preserved(tmp_path):
    pechora = parse(SOURCES / "КС Печорская ГРЭС - Ухта.xlsx")
    assert any(row.control_adp for scheme in pechora.schemes for row in scheme.rows)

    mikun = parse(SOURCES / "Микунь - Сыктывкар.xlsx")
    assert any(row.control_mdp for scheme in mikun.schemes for row in scheme.rows)
    assert any(row.control_mdp_pa for scheme in mikun.schemes for row in scheme.rows)

    syktyvkar = parse(SOURCES / "Сыктывкар.xlsx")
    assert any(
        "[пл]" in item.raw.lower()
        for scheme in syktyvkar.schemes
        for row in scheme.rows
        for item in row.mdp_items
    )


def test_unified_parser_can_export_xlsx(tmp_path):
    source = SOURCES / "КС Печорская ГРЭС - Ухта.xlsx"
    if not source.exists():
        import pytest

        pytest.skip("Sample workbook is not available")
    from mdp_converter.core import convert

    output = tmp_path / "pechora_корр.xlsx"
    model = convert(source, output, output_format="xlsx")
    assert output.exists()
    assert len(model.schemes) > 0

    from openpyxl import load_workbook

    workbook = load_workbook(output)
    assert "Ремонтные схемы" in workbook.sheetnames
    worksheet = workbook["Ремонтные схемы"]
    assert worksheet["A1"].value == "№ п/п"
    assert worksheet["D1"].value == "МДП без ПА"
