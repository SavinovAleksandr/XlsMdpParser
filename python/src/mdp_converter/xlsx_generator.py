"""Write corrected repair-scheme workbooks from a parsed Model."""
from __future__ import annotations

import re
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Model, RowData, Scheme
from .normalization import is_empty_or_dash, low, norm

MINIMAL_FROM = "Минимальный из:"
OUTPUT_SHEET = "Ремонтные схемы"
SUMMARY_SHEET_NAMES = ("Обшая информация о сечении", "Общая информация о сечении", "Информация о сечении")
COLUMN_WIDTHS = (7, 40, 11, 80, 120, 30, 50, 50, 30, 25, 25, 25)
HEADER_FILL = PatternFill("solid", fgColor="B0E0E6")  # PowderBlue
BANNER_FILL = PatternFill("solid", fgColor="FFE4E1")  # MistyRose
THIN_BORDER = Border(
    left=Side(style="thin", color="DFE6EF"),
    right=Side(style="thin", color="DFE6EF"),
    top=Side(style="thin", color="DFE6EF"),
    bottom=Side(style="thin", color="DFE6EF"),
)


def generate_xlsx(model: Model, source_path: str | Path, output_path: str | Path) -> None:
    """Transform a source workbook and save a corrected *_корр-style file."""
    source_path = Path(source_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)

    workbook = load_workbook(output_path)
    _rename_existing_output_sheet(workbook)
    worksheet = workbook.create_sheet(OUTPUT_SHEET, 0)
    flags = _detect_output_flags(model)
    scheme_rows = _write_output_sheet(worksheet, model, flags)
    _update_summary_hyperlinks(workbook, scheme_rows)
    workbook.save(output_path)


def default_xlsx_output_path(source_path: str | Path) -> Path:
    path = Path(source_path)
    stem = path.stem
    if stem.endswith("_корр"):
        return path.with_name(f"{stem}.xlsx")
    return path.with_name(f"{stem}_корр.xlsx")


def _rename_existing_output_sheet(workbook: Workbook) -> None:
    for name in list(workbook.sheetnames):
        if norm(name).casefold() == OUTPUT_SHEET.casefold():
            workbook[name].title = "old"
            return


def _detect_output_flags(model: Model) -> dict[str, bool]:
    has_temperature = any(
        norm(row.temperature)
        for scheme in model.schemes
        for row in scheme.rows
        if not _is_not_controlled(row.temperature)
    )
    has_control_mdp = any(norm(row.control_mdp) for scheme in model.schemes for row in scheme.rows)
    has_control_mdp_pa = any(norm(row.control_mdp_pa) for scheme in model.schemes for row in scheme.rows)
    has_control_adp = any(norm(row.control_adp) for scheme in model.schemes for row in scheme.rows)
    return {
        "temperature": has_temperature,
        "mdp": model.has_mdp,
        "mdp_pa": model.has_mdp_pa,
        "adp": model.has_adp,
        "crit_mdp": model.has_mdp,
        "crit_mdp_pa": model.has_mdp_pa,
        "crit_adp": model.has_adp,
        "ctrl_mdp": has_control_mdp,
        "ctrl_mdp_pa": has_control_mdp_pa,
        "ctrl_adp": has_control_adp,
    }


def _write_output_sheet(worksheet: Worksheet, model: Model, flags: dict[str, bool]) -> dict[str, int]:
    _configure_columns(worksheet)
    _write_headers(worksheet, model.row_axis_label or "ТНВ, °С")
    row_idx = 3
    anchors: dict[str, int] = {}

    for scheme in model.schemes:
        anchor_key = _scheme_anchor_key(scheme.number)
        if anchor_key and anchor_key not in anchors:
            anchors[anchor_key] = row_idx

        banner_text = _scheme_banner_text(scheme.name)
        worksheet.cell(row_idx, 1, scheme.number)
        worksheet.cell(row_idx, 2, banner_text)
        _style_banner_row(worksheet, row_idx, len(COLUMN_WIDTHS))
        row_idx += 1

        body_rows = _scheme_body_rows(scheme)
        if not body_rows:
            continue

        data_start = row_idx
        data_end = data_start + len(body_rows) - 1
        worksheet.cell(data_start, 1, scheme.number)
        worksheet.cell(data_start, 2, scheme.name)
        if data_end > data_start:
            worksheet.merge_cells(start_row=data_start, start_column=1, end_row=data_end, end_column=1)
            worksheet.merge_cells(start_row=data_start, start_column=2, end_row=data_end, end_column=2)

        adp_value = _merged_scheme_value(body_rows, "adp")
        adp_crit = _merged_scheme_value(body_rows, "crit_adp")
        merge_mdp_ctrl = _should_merge_control(body_rows, "control_mdp")
        merge_pa_ctrl = _should_merge_control(body_rows, "control_mdp_pa")
        merge_adp_ctrl = _should_merge_control(body_rows, "control_adp")
        merged_mdp_ctrl = _merged_scheme_value(body_rows, "control_mdp") if merge_mdp_ctrl else ""
        merged_pa_ctrl = _merged_scheme_value(body_rows, "control_mdp_pa") if merge_pa_ctrl else ""
        merged_adp_ctrl = _merged_scheme_value(body_rows, "control_adp") if merge_adp_ctrl else ""

        for offset, row in enumerate(body_rows):
            current = data_start + offset
            if _is_not_controlled(row.temperature):
                worksheet.cell(current, 3, "Не контролируется")
                worksheet.merge_cells(start_row=current, start_column=3, end_row=current, end_column=len(COLUMN_WIDTHS))
                _style_not_controlled(worksheet, current, 3)
                continue

            worksheet.cell(current, 3, row.temperature)
            _set_cell(worksheet, current, 4, _format_mdp_block(row, "mdp", "mdp_items"), flags["mdp"])
            _set_cell(worksheet, current, 5, _format_mdp_block(row, "mdp_pa", "mdp_pa_items"), flags["mdp_pa"])
            if offset == 0 and adp_value and flags["adp"]:
                worksheet.cell(current, 6, adp_value)
                if data_end > data_start:
                    worksheet.merge_cells(start_row=data_start, start_column=6, end_row=data_end, end_column=6)
            _set_cell(worksheet, current, 7, row.crit_mdp, flags["crit_mdp"])
            _set_cell(worksheet, current, 8, row.crit_mdp_pa, flags["crit_mdp_pa"])
            if offset == 0 and adp_crit and flags["crit_adp"]:
                worksheet.cell(current, 9, adp_crit)
                if data_end > data_start:
                    worksheet.merge_cells(start_row=data_start, start_column=9, end_row=data_end, end_column=9)
            if flags["ctrl_mdp"] and not merge_mdp_ctrl:
                _set_cell(worksheet, current, 10, row.control_mdp, True)
            if flags["ctrl_mdp_pa"] and not merge_pa_ctrl:
                _set_cell(worksheet, current, 11, row.control_mdp_pa, True)
            if flags["ctrl_adp"] and not merge_adp_ctrl:
                _set_cell(worksheet, current, 12, row.control_adp, True)

        if merge_mdp_ctrl and merged_mdp_ctrl:
            _apply_vertical_merge(worksheet, data_start, data_end, 10, merged_mdp_ctrl)
        if merge_pa_ctrl and merged_pa_ctrl:
            _apply_vertical_merge(worksheet, data_start, data_end, 11, merged_pa_ctrl)
        if merge_adp_ctrl and merged_adp_ctrl:
            _apply_vertical_merge(worksheet, data_start, data_end, 12, merged_adp_ctrl)

        _style_body_block(worksheet, data_start, data_end, len(COLUMN_WIDTHS))
        row_idx = data_end + 1

    last_row = max(row_idx - 1, 2)
    _apply_table_borders(worksheet, 1, last_row, len(COLUMN_WIDTHS))
    _hide_empty_columns(worksheet, flags)
    worksheet.freeze_panes = "A3"
    return anchors


def _scheme_body_rows(scheme: Scheme) -> list[RowData]:
    if scheme.rows:
        return scheme.rows
    if not scheme.is_controlled:
        return [RowData(temperature="Не контролируется")]
    return []


def _configure_columns(worksheet: Worksheet) -> None:
    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _write_headers(worksheet: Worksheet, row_axis_label: str) -> None:
    headers = (
        ("№ п/п", 1, 1, 2, 1),
        ("Схема сети", 1, 2, 2, 2),
        (row_axis_label if "°" in row_axis_label else f"{row_axis_label}, °С", 1, 3, 2, 3),
        ("МДП без ПА", 1, 4, 2, 4),
        ("МДП с ПА", 1, 5, 2, 5),
        ("АДП", 1, 6, 2, 6),
        ("Критерий определения допустимых перетоков", 1, 7, 1, 9),
        ("Контроль дополнительных параметров", 1, 10, 1, 12),
    )
    subheaders = (
        ("МДП без ПА", 2, 7),
        ("МДП с ПА", 2, 8),
        ("АДП", 2, 9),
        ("МДП без ПА", 2, 10),
        ("МДП с ПА", 2, 11),
        ("АДП", 2, 12),
    )
    for text, r1, c1, r2, c2 in headers:
        worksheet.cell(r1, c1, text)
        if r1 != r2 or c1 != c2:
            worksheet.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for text, row, col in subheaders:
        worksheet.cell(row, col, text)
    for row in (1, 2):
        for col in range(1, len(COLUMN_WIDTHS) + 1):
            cell = worksheet.cell(row, col)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _scheme_banner_text(name: str) -> str:
    return re.sub(r"\s+", " ", norm(name).replace("\n", " "))


def _scheme_anchor_key(number: str) -> str:
    key = norm(number).replace(",", ".")
    while key.endswith("."):
        key = key[:-1]
    return key


def _is_not_controlled(value: str) -> bool:
    return low(value) == "не контролируется"


def _format_mdp_block(row: RowData, raw_attr: str, items_attr: str) -> str:
    raw = norm(getattr(row, raw_attr, ""))
    items = getattr(row, items_attr, [])
    computable = [item for item in items if item.is_computable and norm(item.raw)]
    if len(computable) > 1:
        lines = [MINIMAL_FROM]
        for item in sorted(computable, key=lambda item: item.number):
            lines.append(f"{item.number}) {item.raw}")
        return "\n".join(lines)
    if raw and not low(raw).startswith("минимальн"):
        return raw
    if items:
        lines = []
        for item in sorted(items, key=lambda item: item.number):
            if norm(item.raw):
                lines.append(f"{item.number}) {item.raw}")
        return "\n".join(lines)
    return raw


def _merged_scheme_value(rows: list[RowData], attr: str) -> str:
    for row in rows:
        value = norm(getattr(row, attr, ""))
        if value and not is_empty_or_dash(value):
            return getattr(row, attr, "")
    return ""


def _should_merge_control(rows: list[RowData], attr: str) -> bool:
    values = [norm(getattr(row, attr, "")) for row in rows if norm(getattr(row, attr, ""))]
    unique = list(dict.fromkeys(values))
    return len(unique) == 1


def _set_cell(worksheet: Worksheet, row: int, col: int, value: str, enabled: bool) -> None:
    if not enabled:
        worksheet.cell(row, col, "")
        return
    worksheet.cell(row, col, value or "")
    worksheet.cell(row, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _apply_vertical_merge(worksheet: Worksheet, start_row: int, end_row: int, col: int, value: str) -> None:
    worksheet.cell(start_row, col, value)
    if end_row > start_row:
        worksheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
    worksheet.cell(start_row, col).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _style_banner_row(worksheet: Worksheet, row: int, col_count: int) -> None:
    worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_count)
    for col in range(1, col_count + 1):
        cell = worksheet.cell(row, col)
        cell.fill = BANNER_FILL
        cell.alignment = Alignment(
            horizontal="justify" if col == 2 else "center",
            vertical="center",
            wrap_text=True,
        )


def _style_not_controlled(worksheet: Worksheet, row: int, col: int) -> None:
    cell = worksheet.cell(row, col)
    cell.font = Font(color="FF0000", italic=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_body_block(worksheet: Worksheet, start_row: int, end_row: int, col_count: int) -> None:
    for row in range(start_row, end_row + 1):
        worksheet.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        worksheet.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(4, col_count + 1):
            cell = worksheet.cell(row, col)
            if cell.alignment != Alignment():
                continue
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _apply_table_borders(worksheet: Worksheet, start_row: int, end_row: int, col_count: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            worksheet.cell(row, col).border = THIN_BORDER


def _hide_empty_columns(worksheet: Worksheet, flags: dict[str, bool]) -> None:
    mapping = {
        3: flags["temperature"],
        4: flags["mdp"],
        5: flags["mdp_pa"],
        6: flags["adp"],
        7: flags["crit_mdp"],
        8: flags["crit_mdp_pa"],
        9: flags["crit_adp"],
        10: flags["ctrl_mdp"],
        11: flags["ctrl_mdp_pa"],
        12: flags["ctrl_adp"],
    }
    for col, visible in mapping.items():
        worksheet.column_dimensions[get_column_letter(col)].hidden = not visible


def _update_summary_hyperlinks(workbook: Workbook, anchors: dict[str, int]) -> None:
    if not anchors:
        return
    summary = None
    for name in workbook.sheetnames:
        if any(norm(name).casefold() == norm(candidate).casefold() for candidate in SUMMARY_SHEET_NAMES):
            summary = workbook[name]
            break
    if summary is None:
        return

    link_font = Font(color="0563C1", underline="single")
    for row in summary.iter_rows():
        for cell in row:
            value = norm(cell.value)
            if not value:
                continue
            match = re.match(r"^(\d+(?:\.\d+)?)\s*[.)]?\s*", value.replace(",", "."))
            if not match:
                continue
            key = _scheme_anchor_key(match.group(1))
            target_row = anchors.get(key)
            if target_row is None:
                continue
            cell.hyperlink = f"#{OUTPUT_SHEET}!A{target_row}"
            cell.font = copy(link_font)
